import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

# from openpyxl import Workbook
# Создаем новую рабочую книгу (файл)
# wb = openpyxl.Workbook()
# # Получаем активный лист (по умолчанию создается один лист)
# sheet = wb.active
# # Устанавливаем имя листа
# sheet.title = "Лист1"
# # Записываем данные в ячейки
# save_str = 'B4'
# sheet['A1'] = "Привет"
# sheet['A2'] = "Мир"
# sheet[save_str] = "!!!!!!!!!"
# # Сохраняем файл на диск
# wb.save('test.xlsx')

######################
# октрытие файла xls
######################

# Открываем существующий файл
wb = load_workbook("2_page_test.xlsx")
# Получаем активный лист
sheet = wb.active
# Читаем значение ячейки A1
value = sheet["C1"].value
print(f"Значение ячейки A1: {value}")

contacts = []

#dict_str = [en_word, transc_word_final, transl_word_final, 0, 0]
#dict_all.append(dict_str)  # Включили в состав массива новый блок Слово-Транскрипция-Перевод-Признак учить или нет

i = 1 # первая строка для перебора всех строк файла загрузки. Начинаем с первой строки
while True:
    contact = sheet["A" + str(i)].value  # Контакт (полное ФИО одной строкой)
    post = sheet["B" + str(i)].value     # Должность
    company = sheet["C" + str(i)].value  # Компания
    phone = sheet["D" + str(i)].value    # телефон
    email = sheet["E" + str(i)].value    # e-mail
    second_name = sheet["F" + str(i)].value   # фамилия
    first_name = sheet["G" + str(i)].value    # имя
    patronymic = sheet["H" + str(i)].value    # отчество

    if contact == "" or contact == Truene: # прерываем перебор и считывание как только пойдут пустые строки
        break

    # создали массив одномерный для записи "строки" в большой массив
    new_str = [contact, post, company, phone, email, second_name, first_name, patronymic]

    contacts.append(new_str) # Добавление в массив новой строки-массива
    i+=1


print(contacts)
