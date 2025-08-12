import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

# from openpyxl import Workbook
# Создаем новую рабочую книгу (файл)
wb = openpyxl.Workbook()
# Получаем активный лист (по умолчанию создается один лист)
sheet = wb.active
# Устанавливаем имя листа
sheet.title = "Лист1"
# Записываем данные в ячейки
save_str = 'B4'
sheet['A1'] = "Привет"
sheet['A2'] = "Мир"
sheet[save_str] = "!!!!!!!!!"
# Сохраняем файл на диск
wb.save('test.xlsx')

######################
# октрытие файла xls
######################

# Открываем существующий файл
wb = load_workbook("2_page_test.xlsx")
# Получаем активный лист
sheet = wb.active
# Читаем значение ячейки A1
value = sheet["A1"].value
print(f"Значение ячейки A1: {value}")

